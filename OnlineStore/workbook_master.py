import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.dimensions import DimensionHolder, ColumnDimension

from db_calls import get_category_groups, get_item_groups, get_all_categories, get_all_items, \
    get_all_items_details, get_main_categories_list, get_category_groups_list, get_categories_list, \
    get_items_groups_list, get_category_details_list, get_drop_points_list, get_all_category_details


def check_data_integrity(field_names, obj):
    integrity = True
    for field_name in field_names:
        if obj[field_name] is None:
            integrity = False
            break
    return integrity

def adjust_sheet(ws):
    dim_holder = DimensionHolder(worksheet=ws)

    for col in range(ws.min_column, ws.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=20)
    ws.column_dimensions = dim_holder
    return ws

def create_main_categories_template(data = None):
    template = openpyxl.Workbook()
    working_sheet = template.active

    if data is None:
        working_sheet['A1'] = 'Наименование'
    else:
        working_sheet['A1'] = 'id'
        working_sheet['B1'] = 'Наименование'
        row = 2
        for item in data:
            working_sheet[f'A{row}'] = data[item]
            working_sheet[f'B{row}'] = item
            row +=1
            pass
        pass
    working_sheet = adjust_sheet(working_sheet)
    return template

def create_category_groups_template(data = None):
    template = openpyxl.Workbook()
    working_sheet = template.active
    main_categories_list = get_main_categories_list()
    main_cat_validation = DataValidation(
        type='list', formula1=f'"{",".join(main_categories_list)}"', showDropDown=False
    )

    is_hidden = DataValidation(
        type='list', formula1=','.join(['да', 'нет']), showDropDown=False
    )

    if data is None:
        working_sheet['A1'] = 'Вид товара'
        working_sheet['B1'] = 'Группа Категорий'
        working_sheet['C1'] = 'Скрыто'
        main_cat_validation.add('A2:A10')
        is_hidden.add('C2:C11')
    else:
        working_sheet['A1'] = 'id'
        working_sheet['B1'] = 'Вид товара'
        working_sheet['C1'] = 'Группа Категорий'
        working_sheet['D1'] = 'Скрыто'
        row = 2

        for item in data:
            working_sheet[f'A{row}'] = item['id']
            working_sheet[f'B{row}'] = item['mc_name']
            working_sheet[f'C{row}'] = item['gc_name']
            working_sheet[f'D{row}'] = 'да' if item['gc_is_hidden'] == 1 else 'нет'
            row += 1
            pass
        main_cat_validation.add(f'B2:B{str(len(data) + 1)}')
        is_hidden.add(f'D2:D{str(len(data) + 1)}')
        pass
    working_sheet.add_data_validation(main_cat_validation)
    working_sheet.add_data_validation(is_hidden)
    working_sheet = adjust_sheet(working_sheet)
    return template

def create_item_groups_template(data = None):
    template = openpyxl.Workbook()
    working_sheet = template.active

    is_hidden = DataValidation(
        type='list', formula1=','.join(['да', 'нет']), showDropDown=False
    )

    if data is None:
        working_sheet['A1'] = 'Наименование'
        working_sheet['B1'] = 'Скрыто'
        is_hidden.add('B2:B11')
    else:
        working_sheet['A1'] = 'id'
        working_sheet['B1'] = 'Наименование'
        working_sheet['C1'] = 'Скрыто'
        row = 2

        for group in data:
            working_sheet[f'A{row}'] = group['id']
            working_sheet[f'B{row}'] = group['name']
            working_sheet[f'C{row}'] = 'да' if group['is_hidden'] == 1 else 'нет'
            row += 1
            pass
        is_hidden.add(f'C2:C{row-1}')
        pass
    working_sheet.add_data_validation(is_hidden)
    working_sheet = adjust_sheet(working_sheet)
    return template


def create_categories_template(data = None):
    template = openpyxl.Workbook()
    working_sheet = template.active
    category_group_list = get_category_groups_list()
    escaped_categories = [
        f'"{item}"'
        for item in category_group_list
    ]
    cat_group_validation = DataValidation(
        type='list', formula1=','.join(escaped_categories), showDropDown=False
    )

    is_hidden = DataValidation(
        type='list', formula1=','.join(['да', 'нет']), showDropDown=False
    )

    if data is None:
        working_sheet['A1'] = 'Наименование'
        working_sheet['B1'] = 'Группа категорий'
        working_sheet['C1'] = 'ID альбома'
        working_sheet['D1'] = 'Скрыто'
        cat_group_validation.add('B2:B11')
        is_hidden.add('D2:D11')
        pass
    else:
        working_sheet['A1'] = 'ID'
        working_sheet['B1'] = 'Наименование'
        working_sheet['C1'] = 'Группа категорий'
        working_sheet['D1'] = 'ID альбома'
        working_sheet['E1'] = 'Скрыто'

        row = 2
        for group in data:
            for category in group['categories']:
                working_sheet[f'A{row}'] = category['category_id']
                working_sheet[f'B{row}'] = category['name']
                working_sheet[f'C{row}'] = group['group_name']
                working_sheet[f'D{row}'] = category['icon']
                working_sheet[f'E{row}'] = 'да' if category['category_is_hidden'] == 1 else 'нет'
                row +=1
        cat_group_validation.add(f'C2:C{row-1}')
        is_hidden.add(f'E2:E{row-1}')
    working_sheet.add_data_validation(cat_group_validation)
    working_sheet.add_data_validation(is_hidden)
    working_sheet = adjust_sheet(working_sheet)
    return template

def create_items_template(data = None):
    template = openpyxl.Workbook()
    working_sheet = template.active
    categories_list =[
        category
        for category in get_categories_list()
    ]
    item_groups_list =[
        group
        for group in get_items_groups_list()
    ]

    cat_validation = DataValidation(
        type='list', formula1=','.join(categories_list), showDropDown=False
    )
    item_groups_validation = DataValidation(
        type='list', formula1=','.join(item_groups_list), showDropDown=False
    )

    is_hidden = DataValidation(
        type='list', formula1=','.join(['да', 'нет']), showDropDown=False
    )

    if data is None:
        headers = list(get_category_details_list().keys())
        working_sheet['A1'] = 'Категория'  # список
        working_sheet['B1'] = 'Группа товаров'  # список
        working_sheet['C1'] = 'Название'
        working_sheet['D1'] = 'Цена'
        working_sheet['E1'] = 'Скидка'
        working_sheet['F1'] = 'Альбом'
        working_sheet['G1'] = 'Скрыто'
        col_letter = ord('H')
        for header in headers:
            working_sheet[f'{chr(col_letter)}1'] = header
            col_letter += 1
        cat_validation.add('A2:A11')
        item_groups_validation.add('B2:B11')
        is_hidden.add('G1:G11')
    else:
        headers = data[next(iter(data))]['details']
        working_sheet['A1'] = 'id'
        working_sheet['B1'] = 'Категория' #список
        working_sheet['C1'] = 'Группа товаров' #список
        working_sheet['D1'] = 'Название'
        working_sheet['E1'] = 'Цена'
        working_sheet['F1'] = 'Скидка'
        working_sheet['G1'] = 'Альбом'
        working_sheet['H1'] = 'Скрыто'
        row = 2

        col_letter = ord('I')

        for header in headers:
            working_sheet[f'{chr(col_letter)}1'] = header['name']
            col_letter += 1

        for item_id in data:
            item = data[item_id]
            working_sheet[f'A{row}'] = item['item_id']
            working_sheet[f'B{row}'] = item['category_name']
            working_sheet[f'C{row}'] = item['item_group_name']
            working_sheet[f'D{row}'] = item['name']
            working_sheet[f'E{row}'] = item['price']
            working_sheet[f'F{row}'] = item['discount']
            working_sheet[f'G{row}'] = item['album']
            working_sheet[f'H{row}'] = 'да' if item['is_hidden'] == 1 else 'нет'

            details = item['details']

            col_letter = ord('I')

            for detail in details:
                working_sheet[f'{chr(col_letter)}{row}'] = detail['value']
                col_letter += 1

            row += 1
            cat_validation.add(f'B2:B{row - 1}')
            item_groups_validation.add(f'C2:C{row - 1}')
            is_hidden.add(f'H2:H{row-1}')

    working_sheet.add_data_validation(cat_validation)
    working_sheet.add_data_validation(item_groups_validation)
    working_sheet.add_data_validation(is_hidden)
    working_sheet = adjust_sheet(working_sheet)
    return template

def create_drop_points_template(data = None):
    template = openpyxl.Workbook()
    working_sheet = template.active

    is_hidden = DataValidation(
        type='list', formula1=','.join(['да', 'нет']), showDropDown=False
    )

    if data is None:
        working_sheet['A1'] = 'Страна'
        working_sheet['B1'] = 'Город'
        working_sheet['C1'] = 'Улица'
        working_sheet['D1'] = 'Дом'
        working_sheet['E1'] = 'Примечание'
        working_sheet['F1'] = 'Скрыто'
        is_hidden.add('F1:F11')
    else:
        working_sheet['A1'] = 'id'
        working_sheet['B1'] = 'Страна'
        working_sheet['C1'] = 'Город'
        working_sheet['D1'] = 'Улица'
        working_sheet['E1'] = 'Дом'
        working_sheet['F1'] = 'Примечание'
        working_sheet['G1'] = 'Скрыто'

        row = 2

        for item in data:
            working_sheet[f'A{row}'] = item['drop_point_id']
            working_sheet[f'B{row}'] = item['country']
            working_sheet[f'C{row}'] = item['city']
            working_sheet[f'D{row}'] = item['street']
            working_sheet[f'E{row}'] = item['building']
            working_sheet[f'F{row}'] = item['notes']
            working_sheet[f'G{row}'] = 'да' if item['is_hidden'] == 1 else 'нет'

            row +=1
        is_hidden.add(f'G2:G{row-1}')
    working_sheet.add_data_validation(is_hidden)
    working_sheet = adjust_sheet(working_sheet)
    return template

def create_category_details_template(data=None):
    template = openpyxl.Workbook()
    working_sheet = template.active
    main_categories = get_main_categories_list()

    main_cat_validation = DataValidation(
        type='list', formula1=f'"{",".join(main_categories)}"', showDropDown=False
    )
    is_main_validation = DataValidation(
        type='list', formula1=','.join(['да', 'нет']), showDropDown=False
    )


    if data is None:
        working_sheet['A1'] = 'Главная категория'
        working_sheet['B1'] = 'Наименование'
        working_sheet['C1'] = 'Является главной'
        main_cat_validation.add('A2:A11')
        is_main_validation.add('C2:C11')
        pass
    else:
        row = 2
        working_sheet['A1'] = 'ID'
        working_sheet['B1'] = 'Главная категория'
        working_sheet['C1'] = 'Наименование'
        working_sheet['D1'] = 'Является главной'
        for item in data:
            working_sheet[f'A{row}'] = item['id']
            working_sheet[f'B{row}'] = item['main_category_name']
            working_sheet[f'C{row}'] = item['name']
            working_sheet[f'D{row}'] = 'да' if item['is_main'] == 1 else 'нет'
            row +=1
            pass
        main_cat_validation.add(f'B2:B{row-1}')
        is_main_validation.add(f'D2:D{row-1}')
    working_sheet.add_data_validation(main_cat_validation)
    working_sheet.add_data_validation(is_main_validation)
    working_sheet = adjust_sheet(working_sheet)
    return template

    pass

def get_data_from_file(file_name):
    path = f'storage/{file_name}'
    file = openpyxl.load_workbook(path)
    working_sheet = file.active

    data = []
    for row in working_sheet.iter_rows(min_row=1, min_col=1, max_row=working_sheet.max_row, max_col=working_sheet.max_column):
        row_data = [cell.value for cell in row]
        if row_data[0] == None:
            break
        else:
            data.append(row_data)
    return data

def construct_main_categories(data):
    main_category_template ={
        'id' : None,
        'name' : ''
    }
    main_categories = []

    for i in range(1, len(data)):
        template_copy = main_category_template.copy()
        template_copy['id'] = data[i][0] if len(data[i])>1 else None
        template_copy['name'] = data[i][1] if len(data[i])>1 else data[i][0]
        if not check_data_integrity(['name'], template_copy):
            return None
        else:
            main_categories.append(template_copy)
    return main_categories

def construct_category_groups(data):
    main_categories = get_main_categories_list()
    category_groups = []
    category_template = {
        'id': None,
        'main_category': 0,
        'name': '',
        'is_hidden': 0
    }

    for i in range(1, len(data)):
        template_copy = category_template.copy()
        template_copy['id'] = data[i][0] if len(data[i]) > 3 else None
        template_copy['main_category'] = main_categories[data[i][1]] if len(data[i])>3 else main_categories[data[i][0]]
        template_copy['name'] = data[i][2] if len(data[i]) > 3 else data[i][1]

        is_hidden = data[i][3] if len(data[i]) > 3 else data[i][2]
        template_copy['is_hidden'] = 0 if is_hidden == 'нет' else 1
        if not check_data_integrity(['name', 'main_category', 'is_hidden'], template_copy):
            return None
        else:
            category_groups.append(template_copy)
    return category_groups

def construct_categories(data):
    category_groups_list = get_category_groups_list()

    category_template = {
        'id' : None,
        'name': '',
        'group': 0,
        'album': '',
        'is_hidden': 0
    }
    categories = []
    for row in data[1:]:
        category_template_copy = category_template.copy()
        category_template_copy['id'] = row[0] if len(row)>4 else None
        category_template_copy['name'] = row[1] if len(row) > 4 else row[0]
        category_template_copy['group'] = category_groups_list[row[2]] if len(row) > 4 else category_groups_list[row[1]]
        category_template_copy['album'] = row[3] if len(row) > 4 else row[2]

        is_hidden = row[4] if len(row) > 4 else row[3]
        category_template_copy['is_hidden'] = 0 if is_hidden == 'нет' else 1
        if not check_data_integrity(['name', 'group', 'is_hidden'], category_template_copy):
            return None
        else:
            categories.append(category_template_copy)

    return categories

def construct_item_groups(data):
    category_template = {
        'id': None,
        'name': '',
        'is_hidden' : 0
    }
    categories = []

    for i in range(1, len(data)):
        template_copy = category_template.copy()
        template_copy['id'] = data[i][0] if len(data[i]) > 2 else None
        template_copy['name'] = data[i][1] if len(data[i]) > 2 else data[i][0]

        is_hidden = data[i][2] if len(data[i]) > 2 else data[i][1]
        template_copy['is_hidden'] = 0 if is_hidden == 'нет' else 1

        if not check_data_integrity(['name', 'is_hidden'], template_copy):
            return None
        else:
            categories.append(template_copy)
    return categories

def construct_items_edit(data):
    item_template = {
        'id': None,
        'category': 0,
        'group': 0,
        'name': '',
        'price': 0,
        'discount': 0,
        'album': '',
        'is_hidden': 0,
        'details': []
    }
    categories_list = get_categories_list()
    category_details_list = get_category_details_list()
    item_groups_list = get_items_groups_list()

    items_main_data = []
    for i in range(1, len(data)):
        item_details = []
        item_template_copy = item_template.copy()
        for j, item_detail in enumerate(item_template_copy):
            if item_detail == 'category':
                item_template_copy[item_detail] = categories_list[f'"{data[i][j]}"']
            elif item_detail == 'group':
                item_template_copy[item_detail] = item_groups_list[f'"{data[i][j]}"']
            elif item_detail == 'is_hidden':
                item_template_copy[item_detail] = 0 if data[i][j] == 'нет' else 1
            elif item_detail == 'details':
                for l in range(j, len(data[i])):
                    if data[i][l] is not None:
                        item_details.append({
                            'id': category_details_list[data[0][l]],
                            'value': data[i][l]
                        })
            else:
                item_template_copy[item_detail]=data[i][j]

        item_template_copy['details'] = item_details
        if not check_data_integrity(['category', 'group', 'name', 'price', 'is_hidden'], item_template_copy):
            return None
        else:
            items_main_data.append(item_template_copy)

    return items_main_data


def construct_items_new(data):
    item_template = {
        'category': 0,
        'group': 0,
        'name': '',
        'price': 0,
        'discount': 0,
        'album': '',
        'is_hidden': 0,
        'details': []
    }
    categories_list = get_categories_list()
    category_details_list = get_category_details_list()
    item_groups_list = get_items_groups_list()

    items_main_data = []
    for i in range(1, len(data)):
        item_details = []
        item_template_copy = item_template.copy()
        for j, item_detail in enumerate(item_template_copy):

            if item_detail == 'category':
                item_template_copy[item_detail] = categories_list[f'"{data[i][j]}"']
            elif item_detail == 'group':
                item_template_copy[item_detail] = item_groups_list[f'"{data[i][j]}"']
            elif item_detail == 'is_hidden':
                item_template_copy[item_detail] = 0 if data[i][j] == 'нет' else 1
            elif item_detail == 'details':
                for l in range(j, len(data[i])):
                    if data[i][l] is not None:
                        item_details.append({
                            'id': category_details_list[data[0][l]],
                            'value': data[i][l]
                        })
            else:
                item_template_copy[item_detail]=data[i][j]

        item_template_copy['details'] = item_details
        if not check_data_integrity(['category', 'group', 'name', 'price', 'is_hidden'], item_template_copy):
            return None
        else:
            items_main_data.append(item_template_copy)

    return items_main_data

def construct_items(action, data):
    if action == 'add':
        return construct_items_new(data)
    else:
        return construct_items_edit(data)

def construct_drop_point(data):
    dp_template = {
        'id' : None,
        'country' : '',
        'city' : '',
        'street': '',
        'building' : '',
        'notes' : '',
        'is_hidden' : 0
    }

    drop_points = []

    for i in range(1, len(data)):
        dp_template_copy = dp_template.copy()
        dp_template_copy['id'] = data[i][0] if len(data[i]) > 6 else None
        dp_template_copy['country'] = data[i][1] if len(data[i]) > 6 else data[i][0]
        dp_template_copy['city'] = data[i][2] if len(data[i]) > 6 else data[i][1]
        dp_template_copy['street'] = data[i][3] if len(data[i]) > 6 else data[i][2]
        dp_template_copy['building'] = data[i][4] if len(data[i]) > 6 else data[i][3]
        dp_template_copy['notes'] = data[i][5] if len(data[i]) > 6 else data[i][4]

        is_hidden = data[i][6] if len(data[i]) > 6 else data[i][5]
        dp_template_copy['is_hidden'] = 0 if is_hidden == 'нет' else 1

        if not check_data_integrity(['country', 'city', 'street', 'building', 'is_hidden'], dp_template_copy):
            return None
        else:
            drop_points.append(dp_template_copy)
    return drop_points

def construct_category_details(data):
    category_detail_template = {
        'id' : None,
        'main_category_id': None,
        'name' : '',
        'is_main': False
    }
    main_categories_list = get_main_categories_list()
    category_details = []
    for i in range(1, len(data)):
        cd_template_copy = category_detail_template.copy()
        cd_template_copy['id'] = data[i][0] if len(data[i]) > 3 else None
        cd_template_copy['main_category_id'] = main_categories_list[data[i][1]] if len(data[i]) > 3 else main_categories_list[data[i][0]]
        cd_template_copy['name'] = data[i][2] if len(data[i]) > 3 else data[i][1]
        if len(data[i]) > 3:
            cd_template_copy['is_main'] = 1 if data[i][3] == 'да' else 0
        else:
            cd_template_copy['is_main'] = 1 if data[i][2] == 'да' else 0

        if not check_data_integrity(['main_category_id', 'name','is_main'], cd_template_copy):
            return None

        else:
            category_details.append(cd_template_copy)
    return category_details

def save_doc(doc, mod):
    doc.save(f'storage/doc_{mod}.xlsx')
